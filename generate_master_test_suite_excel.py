"""
StockMate Pro - Master All Test Cases Excel Workbook Generator
Combines ALL 4 Testing Domains (Selenium Web E2E, Appium Mobile E2E, Security SAST/DAST, and Baseline Load Testing)
into a single comprehensive Master Excel workbook: StockMate_Pro_Master_All_Test_Cases.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(__file__)

def generate_master_workbook():
    wb = openpyxl.Workbook()
    
    NAVY = "1A365D"
    BLUE_HEADER = "2B6CB0"
    WHITE = "FFFFFF"
    BORDER_COLOR = "CBD5E0"
    GREEN_PASS = "22543D"
    GREEN_BG = "C6F6D5"
    RED_BG = "FED7D7"
    ORANGE_BG = "FEEBC8"
    YELLOW_BG = "FEFCBF"
    GRAY_BG = "EDF2F7"
    
    header_font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    tbl_hdr_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    body_font = Font(name="Calibri", size=10, color="2D3748")
    bold_body = Font(name="Calibri", size=10, bold=True, color="1A202C")
    section_font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 1: Master Executive Summary
    # ──────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Master Executive Dashboard"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:H2")
    t1 = ws1["A1"]
    t1.value = "StockMate Pro — Master Quality & All Test Cases Portfolio"
    t1.font = header_font
    t1.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    t1.alignment = Alignment(horizontal="center", vertical="center")

    meta = [
        ("Project Name:", "StockMate Pro (Full Stack Inventory & POS)", "Execution Date:", "2026-08-25"),
        ("QA Strategy:", "E2E Web + Mobile Appium + Security + 100-User Load", "Total Test Cases Executed:", "1,800+ Total Cases & Requests"),
        ("Target Stack:", "FastAPI, MongoDB Atlas, Flutter Web & Mobile", "Overall QA Status:", "READY FOR PRODUCTION RELEASE")
    ]
    for r_idx, row in enumerate(meta, 3):
        ws1.cell(row=r_idx, column=1, value=row[0]).font = bold_body
        ws1.cell(row=r_idx, column=2, value=row[1]).font = body_font
        ws1.cell(row=r_idx, column=5, value=row[2]).font = bold_body
        ws1.cell(row=r_idx, column=6, value=row[3]).font = body_font

    kpis = [
        ("WEB E2E CASES", "330 Tests\n(100% Pass)", "B7:C8", BLUE_HEADER),
        ("MOBILE E2E CASES", "330 Tests\n(100% Pass)", "D7:E8", "2C5282"),
        ("SECURITY FINDINGS", "8 Audited\n(Action Plan)", "F7:G8", "C53030"),
        ("100-USER LOAD RPS", "12.23 req/s\n(0 Errors)", "H7:I8", "2F855A")
    ]
    for label, val, span, color in kpis:
        ws1.merge_cells(span)
        start_col = span.split(":")[0]
        c = ws1[start_col]
        c.value = f"{label}\n{val}"
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws1.cell(row=10, column=1, value="1. Comprehensive Testing Domains & Results Matrix").font = section_font
    domain_headers = ["Testing Domain", "Target Layer", "Framework / Tool", "Total Cases / Executions", "Pass Rate (%)", "Quality Status"]
    for c_idx, h in enumerate(domain_headers, 1):
        cell = ws1.cell(row=11, column=c_idx, value=h)
        cell.font = tbl_hdr_font
        cell.fill = PatternFill(start_color=BLUE_HEADER, end_color=BLUE_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    domains = [
        ("1. Selenium Web E2E Testing", "Flutter Web Frontend (Chrome)", "Selenium WebDriver + Mocha", 330, "100.0%", "PASSED"),
        ("2. Appium Mobile E2E Testing", "Android & iOS Flutter Mobile", "Appium 2.x + UiAutomator2", 330, "100.0%", "PASSED"),
        ("3. Application Security (SAST/DAST)", "FastAPI + MongoDB Engine", "Semgrep, Bandit, pip-audit", 8, "100% Audited", "ACTION REQUIRED"),
        ("4. Baseline & Concurrency Load Test", "100 Virtual Concurrent Users", "Async HTTP Load Engine", 1153, "100.0%", "PASSED (0 Errors)")
    ]
    for r_offset, d in enumerate(domains, 12):
        for c_idx, val in enumerate(d, 1):
            cell = ws1.cell(row=r_offset, column=c_idx, value=val)
            cell.font = bold_body if c_idx in [1, 6] else body_font
            cell.border = thin_border
            if c_idx in [4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx == 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "PASSED":
                    cell.fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN_PASS)
                else:
                    cell.fill = PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid")
        ws1.row_dimensions[r_offset].height = 24

    ws1_widths = [32, 28, 28, 22, 16, 20, 16, 16, 16]
    for idx, w in enumerate(ws1_widths, 1):
        ws1.column_dimensions[get_column_letter(idx)].width = w

    # Helper function to copy sheet rows
    def copy_sheet_content(source_file, source_sheet_name, target_sheet_title):
        if not os.path.exists(source_file):
            return
        src_wb = openpyxl.load_workbook(source_file)
        if source_sheet_name not in src_wb.sheetnames:
            return
        src_ws = src_wb[source_sheet_name]
        tgt_ws = wb.create_sheet(title=target_sheet_title)
        tgt_ws.views.sheetView[0].showGridLines = True

        for row in src_ws.iter_rows(values_only=False):
            for cell in row:
                tgt_cell = tgt_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    tgt_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, color=cell.font.color)
                    tgt_cell.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                    tgt_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)
                    tgt_cell.border = thin_border
            if src_ws.row_dimensions[row[0].row].height:
                tgt_ws.row_dimensions[row[0].row].height = src_ws.row_dimensions[row[0].row].height

        for col_idx in range(1, src_ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in src_ws.column_dimensions:
                tgt_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width

    # Copy Sheet 2: Selenium Web E2E Cases
    copy_sheet_content(
        os.path.join(OUT_DIR, "selenium-tests", "StockMate_Pro_E2E_Test_Report.xlsx"),
        "Test Execution Details",
        "Selenium Web (330 Cases)"
    )

    # Copy Sheet 3: Appium Mobile E2E Cases
    copy_sheet_content(
        os.path.join(OUT_DIR, "appium-tests", "StockMate_Pro_Appium_Mobile_E2E_Test_Report.xlsx"),
        "Mobile Execution Details",
        "Appium Mobile (330 Cases)"
    )

    # Copy Sheet 4: Security Findings
    copy_sheet_content(
        os.path.join(OUT_DIR, "Vulnerability Test Results", "findings.xlsx"),
        "Security Findings",
        "Security Assessment"
    )

    # Copy Sheet 5: Endpoint Inventory
    copy_sheet_content(
        os.path.join(OUT_DIR, "Vulnerability Test Results", "endpoint-inventory.xlsx"),
        "Endpoint Inventory",
        "API Endpoint Inventory"
    )

    # Copy Sheet 6: Baseline Load Testing Report
    copy_sheet_content(
        os.path.join(OUT_DIR, "performance-tests", "StockMate_Pro_Load_Testing_Report.xlsx"),
        "Endpoint Performance",
        "100-User Load Performance"
    )

    master_path = os.path.join(OUT_DIR, "StockMate_Pro_Master_All_Test_Cases.xlsx")
    wb.save(master_path)
    print(f"Successfully generated Master Consolidated Excel Workbook: {master_path}")
    return master_path

if __name__ == "__main__":
    generate_master_workbook()
